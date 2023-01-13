# Importing the required libraries from selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from fake_useragent import UserAgent
from selenium.webdriver.support import expected_conditions as ec
from datetime import datetime as dt
from time import sleep
import numpy as np
import warnings
warnings.filterwarnings("ignore")
import glob
import shutil
import pandas as pd
from credentials import *

# Library for checking time of execution
import time

# Library for working with excel files
import openpyxl

# Importing library for auto installing the correct version of chrome driver needed to open the browser
import chromedriver_autoinstaller
import os

def Scrape_slider_info():
    # Finding the elements for memory sizes of the phone
    slider = driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
    info = slider[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

     # Making list for all memory sizes of the phone
    slider_info = []
    for inf in info:
        slider_info.append(inf)
        
    return slider_info

def Scrape_name_and_click_through(counter):
    # Finding the elements for memory sizes of the phone (last element)
    slider = driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
    info_n = slider[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')[counter]

    # Getting the name of the slider n-th element
    name = info_n.text

    # Clicking to continue to next combination
    return name, info_n

def Go_back():
    # Going back for memory size options
    go_back = driver.find_elements_by_xpath("//button[@class='btnAnswer']")
    go_back[-1].click() 
    
def Get_price():
    price = driver.find_element_by_xpath('//div[@class="animated bounceIn"]').text
    
    return price

def Print_info(phone_url, phone_name, gen_name, size_name, display_name, frame_name, back_name, price):
    print("****************************************************")
    print("Source URL: ", phone_url)
    print("Iphone Model: ", phone_name)
    print("Phone Condition: ", gen_name)
    print("Memory Size: ", size_name)
    print("Display Condition: ", display_name)
    print("Frame Condition: ", frame_name)
    print("Back Condition: ", back_name)
    print("Price: ", price)
    
def Store_info(phone_url, phone_name, gen_name, size_name, display_name, frame_name, back_name, price, counter):
    c1 = sheet1.cell(row=counter, column=1)
    c1.value = phone_url
    c2 = sheet1.cell(row=counter, column=2)
    c2.value = phone_name
    c3 = sheet1.cell(row=counter, column=3)
    c3.value = gen_name
    c4 = sheet1.cell(row=counter, column=4)
    c4.value = size_name
    c5 = sheet1.cell(row=counter, column=5)
    c5.value = display_name
    c6 = sheet1.cell(row=counter, column=6)
    c6.value = frame_name
    c7 = sheet1.cell(row=counter, column=7)
    c7.value = back_name
    c8 = sheet1.cell(row=counter, column=8)
    c8.value = price


start_time = time.time()

# Creating new workbook in order to store the infomation from the page
excel = openpyxl.Workbook()
# Activating the workbook
sheet1 = excel.active
sheet1.title = website_url
# Creaiting new columns with headers for every category
c1 = sheet1.cell(row=1, column=1)
c1.value = "SOURCE_URL"
c2 = sheet1.cell(row=1, column=2)
c2.value = "Phone"
c3 = sheet1.cell(row=1, column=3)
c3.value = "Phone Condition"
c4 = sheet1.cell(row=1, column=4)
c4.value = "Memory Size"
c5 = sheet1.cell(row=1, column=5)
c5.value = "Display Condition"
c6 = sheet1.cell(row=1, column=6)
c6.value = "Frame Condition"
c7 = sheet1.cell(row=1, column=7)
c7.value = "Back Condition"
c8 = sheet1.cell(row=1, column=8)
c8.value = "Price"
# Adding counter in order to iterate through new rows and store the information

# Function for scraping information from the website only for iphones
def Scraping(driver, counter_fails = 0, index = 0, counter = 2):

    driver.get(website_url)

    sleep(5)
    samsung_type=driver.find_elements_by_xpath("//a[@class='verkaufenPageAnswersBeta modelAnswer']")

    # Making list of all iphone models
    list_phones_url = []
    phone_names = []
    for phone in samsung_type:
        if 'Galaxy S' in phone.text or 'Galaxy A' in phone.text:
            phone_names.append(phone.text)
            list_phones_url.append((phone.get_attribute('href'), phone.text))
            
    # print(phone_names.index('Galaxy S10 Lite'))
    
    try:
        for phone in list_phones_url[index:]:
            if 'S9' in phone[1] or 'S8' in phone[1] or 'S7' in phone[1]:
                pass
            else:
                index_at = list_phones_url.index(phone)
                print(phone[1])

                # Getting the url
                phone_url = phone[0]

                # Getting the phone name
                phone_name = phone[1]

                # Navigating to the page
                driver.get(phone_url)

                sleep(4)
                # Finding the elements for general condition of the phone
                gen_type=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                gen_a = gen_type[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                # Making list for all types of phone conditions and for the names
                list_gen_name = []
                list_gen = []
                for gen in gen_a:
                    list_gen.append(gen)

                # Going only 2 iterations in order to skip Defekt conditions
                for i in range(2):
                    sleep(1.5)

                    # Finding the elements for general condition of the phone (last element)
                    gen_type=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                    gen_n = gen_type[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')[i]

                    # Getting the name of the condition
                    gen_name = gen_n.text

                    if gen_name in list_gen_name:
                        continue
                    else:
                        list_gen_name.append(gen_name)
                        # Clicking to continue to next combination
                        gen_n.click()

                        sleep(1.5)

                        if phone_name == 'Galaxy S21 Ultra' or phone_name == 'Galaxy S20 Ultra 5G':
                            # Finding the elements for the colors of the phone
                            colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                            colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                            # Clicking to continue to next combination
                            colors[0].click()
                            time.sleep(1.5)

                            mem_sizes = Scrape_slider_info()
                            list_size_name = []

                            for i in range(len(mem_sizes)):
                                name_size, size_n = Scrape_name_and_click_through(i)

                                if name_size in list_size_name:
                                    continue
                                else:
                                    list_size_name.append(name_size)

                                size_n.click()

                                time.sleep(1.5)

                                # Checking whether the condition is 'Second hand' because there is different elements
                                if gen_name == 'Gebraucht':
                                    dis_conds = Scrape_slider_info()
                                    list_dis_cond = []

                                    for i in range(len(dis_conds)):
                                        name_dis, dis_n = Scrape_name_and_click_through(i)

                                        if name_dis in list_dis_cond:
                                            continue
                                        else:
                                            list_dis_cond.append(name_dis)

                                        dis_n.click()

                                        time.sleep(1.5)

                                        # Finding the elements for the frame condition of the phone
                                        frame_conds = Scrape_slider_info()
                                        list_frame_cond = []

                                        for i in range(len(frame_conds)):
                                            name_frame, frame_n = Scrape_name_and_click_through(i)

                                            if name_frame in list_frame_cond:
                                                continue
                                            else:
                                                list_frame_cond.append(name_frame)

                                            frame_n.click()

                                            time.sleep(1.5)

                                            # Finding the elements for the back condition of the phone
                                            back_conds = Scrape_slider_info()
                                            list_back_cond = []

                                            for i in range(len(back_conds)):
                                                name_back, back_n = Scrape_name_and_click_through(i)

                                                if name_back in list_back_cond:
                                                    continue
                                                else:
                                                    list_back_cond.append(name_back)

                                                back_n.click()

                                                time.sleep(1.5)

                                                price = Get_price()

                                                Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                                Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                                counter += 1

                                                Go_back()

                                            Go_back()

                                        Go_back()

                                # Checking whether the phone condition is Poor or Defekt in order to skip
                                elif gen_name == 'Mangelhaft / Defekt':
                                    print("Passing Condition", gen_name)
                                    pass

                                else:
                                    time.sleep(1.5)
                                    price = Get_price()

                                    name_dis = np.NaN
                                    name_frame = np.NaN
                                    name_back = np.NaN

                                    Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                    Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                    counter += 1

                                Go_back()

                            Go_back()

                        elif phone_name == 'Galaxy S10 Lite' or phone_name == 'Galaxy S20+ 5G' or phone_name == 'Galaxy S20+' or phone_name == 'Galaxy S20 5G' or phone_name == 'Galaxy S20 4G' or phone_name == 'Galaxy A20e (2019)':  
                            # Checking whether the condition is 'Second hand' because there is different elements
                            if gen_name == 'Gebraucht':
                                name_size = np.NaN
                                # Finding the elements for the colors of the phone
                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                colors[0].click()
                                time.sleep(1.5)

                                dis_conds = Scrape_slider_info()
                                list_dis_cond = []

                                for i in range(len(dis_conds)):
                                    name_dis, dis_n = Scrape_name_and_click_through(i)

                                    if name_dis in list_dis_cond:
                                        continue
                                    else:
                                        list_dis_cond.append(name_dis)

                                    dis_n.click()
                                    time.sleep(1.5)

                                    # Finding the elements for the frame condition of the phone
                                    frame_conds = Scrape_slider_info()
                                    list_frame_cond = []

                                    for i in range(len(frame_conds)):
                                        name_frame, frame_n = Scrape_name_and_click_through(i)

                                        if name_frame in list_frame_cond:
                                            continue
                                        else:
                                            list_frame_cond.append(name_frame)

                                        frame_n.click()
                                        time.sleep(1.5)

                                        # Finding the elements for the back condition of the phone
                                        back_conds = Scrape_slider_info()
                                        list_back_cond = []

                                        for i in range(len(back_conds)):
                                            name_back, back_n = Scrape_name_and_click_through(i)

                                            if name_back in list_back_cond:
                                                continue
                                            else:
                                                list_back_cond.append(name_back)

                                            back_n.click()

                                            time.sleep(1.5)

                                            price = Get_price()

                                            Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                            Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                            counter += 1

                                            Go_back()

                                        Go_back()

                                    Go_back()

                                Go_back()
                                time.sleep(0.5)
                            # Checking whether the phone condition is Poor or Defekt in order to skip
                            elif gen_name == 'Mangelhaft / Defekt':
                                print("Passing Condition", gen_name)
                                pass

                            # If the phone condition is New (Welded) then execute this block
                            else:
                                # Finding the elements for the colors of the phone
                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                colors[0].click()
                                sleep(1.5)

                                price = Get_price()

                                name_dis = np.NaN
                                name_frame = np.NaN
                                name_back = np.NaN
                                name_size = np.NaN

                                Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                counter += 1

                                # Going back for memory size options
                                Go_back()

                        elif phone_name == 'Galaxy S20 FE':
                            name_size = '128GB'

                            # Checking whether the condition is 'Second hand' because there is different elements
                            if gen_name == 'Gebraucht':
                                memory=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                mem = memory[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                mem[0].click()
                                time.sleep(1)

                                # Finding the elements for the colors of the phone
                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                colors[0].click()
                                time.sleep(1.5)

                                dis_conds = Scrape_slider_info()
                                list_dis_cond = []

                                for i in range(len(dis_conds)):
                                    name_dis, dis_n = Scrape_name_and_click_through(i)

                                    if name_dis in list_dis_cond:
                                        continue
                                    else:
                                        list_dis_cond.append(name_dis)

                                    dis_n.click()

                                    time.sleep(1.5)

                                    # Finding the elements for the frame condition of the phone
                                    frame_conds = Scrape_slider_info()
                                    list_frame_cond = []

                                    for i in range(len(frame_conds)):
                                        name_frame, frame_n = Scrape_name_and_click_through(i)

                                        if name_frame in list_frame_cond:
                                            continue
                                        else:
                                            list_frame_cond.append(name_frame)

                                        frame_n.click()

                                        time.sleep(1.5)

                                        # Finding the elements for the back condition of the phone
                                        back_conds = Scrape_slider_info()
                                        list_back_cond = []

                                        for i in range(len(back_conds)):
                                            name_back, back_n = Scrape_name_and_click_through(i)

                                            if name_back in list_back_cond:
                                                continue
                                            else:
                                                list_back_cond.append(name_back)

                                            back_n.click()

                                            time.sleep(1.5)

                                            # Finding the elements for the colors of the phone
                                            dual_sim=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                            sims = dual_sim[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                            # Clicking to continue to next combination
                                            sims[0].click()
                                            time.sleep(1.5)

                                            price = Get_price()

                                            Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                            Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                            counter += 1

                                            Go_back()

                                            Go_back()

                                        Go_back()

                                    Go_back()
                            # Checking whether the phone condition is Poor or Defekt in order to skip
                            elif gen_name == 'Mangelhaft / Defekt':
                                print("Passing Condition", gen_name)
                                pass

                            # If the phone condition is New (Welded) then execute this block
                            else:
                                # Finding the elements for the colors of the phone
                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                colors[0].click()
                                sleep(1.5)
                                # Finding the elements for the colors of the phone
                                dual_sim=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                sims = dual_sim[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                sims[0].click()
                                sleep(1.5)

                                price = Get_price()

                                name_dis = np.NaN
                                name_frame = np.NaN
                                name_back = np.NaN

                                Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                counter += 1

                                # Going back for memory size options
                                Go_back()

                            # Going back for memory size options
                            Go_back()

                            # Going back for memory size options
                            Go_back()

                        elif phone_name == 'Galaxy S20 FE 5G' or phone_name == 'Galaxy S10 5G' or phone_name == 'Galaxy A51 5G (2020)' or phone_name == 'Galaxy A40 (2019)' or phone_name == 'Galaxy A51 (2020)' or phone_name == 'Galaxy A70 (2019)' or phone_name == 'Galaxy A51 5G (2020)' or phone_name == 'Galaxy A41 (2020)' or phone_name == 'Galaxy A21s (2020)' or phone_name == 'Galaxy A20s' or phone_name == 'Galaxy A42 5G (2020)':
                            if phone_name == 'Galaxy S20 FE 5G' or phone_name == 'Galaxy A42 5G (2020)':
                                name_size = '128GB'
                            elif phone_name == 'Galaxy A40 (2019)':
                                name_size = '64GB'
                            elif phone_name == 'Galaxy A70 (2019)' or phone_name == 'Galaxy A51 (2020)' or phone_name == 'Galaxy A51 5G (2020)' or phone_name == 'Galaxy A41 (2020)' or phone_name == 'Galaxy A21s (2020)':
                                name_size = np.NaN
                            elif phone_name == 'Galaxy A20s':
                                name_size = '32GB'
                            else:
                                name_size = '256GB'

                            # Checking whether the condition is 'Second hand' because there is different elements
                            if gen_name == 'Gebraucht':
                                if phone_name == 'Galaxy A40 (2019)' or phone_name == 'Galaxy A70 (2019)' or phone_name == 'Galaxy A51 (2020)' or phone_name == 'Galaxy A41 (2020)' or phone_name == 'Galaxy A21s (2020)' or phone_name == 'Galaxy A20s' or phone_name == 'Galaxy A42 5G (2020)' or phone_name == 'Galaxy S10 5G':
                                    pass
                                else:
                                    memory=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                    mem = memory[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                    # Clicking to continue to next combination
                                    mem[0].click()
                                    time.sleep(1)

                                # Finding the elements for the colors of the phone
                                time.sleep(1)
                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                colors[0].click()
                                time.sleep(1.5)

                                dis_conds = Scrape_slider_info()
                                list_dis_cond = []

                                for i in range(len(dis_conds)):
                                    name_dis, dis_n = Scrape_name_and_click_through(i)

                                    if name_dis in list_dis_cond:
                                        continue
                                    else:
                                        list_dis_cond.append(name_dis)

                                    dis_n.click()

                                    time.sleep(1.5)

                                    # Finding the elements for the frame condition of the phone
                                    frame_conds = Scrape_slider_info()
                                    list_frame_cond = []

                                    for i in range(len(frame_conds)):
                                        name_frame, frame_n = Scrape_name_and_click_through(i)

                                        if name_frame in list_frame_cond:
                                            continue
                                        else:
                                            list_frame_cond.append(name_frame)

                                        frame_n.click()

                                        time.sleep(1.5)

                                        # Finding the elements for the back condition of the phone
                                        back_conds = Scrape_slider_info()
                                        list_back_cond = []

                                        for i in range(len(back_conds)):
                                            name_back, back_n = Scrape_name_and_click_through(i)

                                            if name_back in list_back_cond:
                                                continue
                                            else:
                                                list_back_cond.append(name_back)

                                            back_n.click()

                                            time.sleep(1.5)

                                            price = Get_price()

                                            Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                            Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                            counter += 1

                                            # Going back for back conditions
                                            Go_back()

                                        # Going back for frame conditions
                                        Go_back()

                                    # Going back for display conditions
                                    Go_back()

                            # Checking whether the phone condition is Poor or Defekt in order to skip
                            elif gen_name == 'Mangelhaft / Defekt':
                                print("Passing Condition", gen_name)
                                pass

                            # If the phone condition is New (Welded) then execute this block
                            else:
                                # Finding the elements for the colors of the phone
                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                colors[0].click()
                                sleep(1.5)

                                price = Get_price()

                                name_dis = np.NaN
                                name_frame = np.NaN
                                name_back = np.NaN

                                Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                counter += 1

                            # Going back for color options
                            if phone_name == 'Galaxy A70 (2019)' or phone_name == 'Galaxy A51 (2020)' or phone_name == 'Galaxy A41 (2020)' or phone_name == 'Galaxy A21s (2020)':
                                pass
                            else:
                                Go_back()
                                time.sleep(0.5)

                            # Going back for memory size options
                            Go_back()

                        elif phone_name == 'Galaxy S10+' or phone_name == 'Galaxy S10':
                            # Finding the elements for memory sizes of the phone
                            mem_sizes = Scrape_slider_info()
                            list_size_name = []

                            for i in range(len(mem_sizes)):
                                name_size, size_n = Scrape_name_and_click_through(i)

                                if name_size in list_size_name:
                                    continue
                                else:
                                    list_size_name.append(name_size)

                                size_n.click()

                                time.sleep(1.5)

                                dual_sim=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                sims = dual_sim[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                sims[0].click()
                                time.sleep(1)

                                # Finding the elements for the colors of the phone
                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                colors[0].click()
                                time.sleep(1)

                                # Checking whether the condition is 'Second hand' because there is different elements
                                if gen_name == 'Gebraucht':
                                    dis_conds = Scrape_slider_info()
                                    list_dis_cond = []

                                    for i in range(len(dis_conds)):
                                        name_dis, dis_n = Scrape_name_and_click_through(i)

                                        if name_dis in list_dis_cond:
                                            continue
                                        else:
                                            list_dis_cond.append(name_dis)

                                        dis_n.click()

                                        time.sleep(1.5)

                                        # Finding the elements for the frame condition of the phone
                                        frame_conds = Scrape_slider_info()
                                        list_frame_cond = []

                                        for i in range(len(frame_conds)):
                                            name_frame, frame_n = Scrape_name_and_click_through(i)

                                            if name_frame in list_frame_cond:
                                                continue
                                            else:
                                                list_frame_cond.append(name_frame)

                                            frame_n.click()

                                            time.sleep(1.5)

                                            # Finding the elements for the back condition of the phone
                                            back_conds = Scrape_slider_info()
                                            list_back_cond = []

                                            for i in range(len(back_conds)):
                                                name_back, back_n = Scrape_name_and_click_through(i)

                                                if name_back in list_back_cond:
                                                    continue
                                                else:
                                                    list_back_cond.append(name_back)

                                                back_n.click()

                                                time.sleep(1.5)

                                                price = Get_price()

                                                Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                                Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                                counter += 1

                                                Go_back()

                                            Go_back()

                                        Go_back()

                                # Checking whether the phone condition is Poor or Defekt in order to skip
                                elif gen_name == 'Mangelhaft / Defekt':
                                    print("Passing Condition", gen_name)
                                    pass


                                # Going back for memory size options
                                Go_back()
                                time.sleep(0.5)

                                # Going back for memory size options
                                Go_back()
                                time.sleep(0.5)

                                # Going back for color options
                                Go_back()

                        elif phone_name == 'Galaxy S10e' or phone_name == 'Galaxy A50 (2019)':
                            # Checking whether the condition is 'Second hand' because there is different elements
                            if gen_name == 'Gebraucht':
                                if phone_name == 'Galaxy A50 (2019)':
                                    pass
                                else:
                                    dual_sim=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                    sims = dual_sim[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                    # Clicking to continue to next combination
                                    sims[0].click()
                                    time.sleep(1)

                                # Finding the elements for the colors of the phone
                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                colors[0].click()
                                time.sleep(1)

                                # Finding the elements for memory sizes of the phone
                                mem_sizes = Scrape_slider_info()
                                list_size_name = []

                                for i in range(len(mem_sizes)):
                                    name_size, size_n = Scrape_name_and_click_through(i)

                                    if name_size in list_size_name:
                                        continue
                                    else:
                                        list_size_name.append(name_size)

                                    size_n.click()

                                    time.sleep(1.5)

                                    if phone_name == 'Galaxy A50 (2019)':
                                        # Finding the elements for the GB memory
                                        memory_gb=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                        memory = memory_gb[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                        # Clicking to continue to next combination
                                        memory[0].click()
                                        time.sleep(1)

                                    dis_conds = Scrape_slider_info()
                                    list_dis_cond = []

                                    for i in range(len(dis_conds)):
                                        name_dis, dis_n = Scrape_name_and_click_through(i)

                                        if name_dis in list_dis_cond:
                                            continue
                                        else:
                                            list_dis_cond.append(name_dis)

                                        dis_n.click()

                                        time.sleep(1.5)

                                        # Finding the elements for the frame condition of the phone
                                        frame_conds = Scrape_slider_info()
                                        list_frame_cond = []

                                        for i in range(len(frame_conds)):
                                            name_frame, frame_n = Scrape_name_and_click_through(i)

                                            if name_frame in list_frame_cond:
                                                continue
                                            else:
                                                list_frame_cond.append(name_frame)

                                            frame_n.click()

                                            time.sleep(1.5)

                                            # Finding the elements for the back condition of the phone
                                            back_conds = Scrape_slider_info()
                                            list_back_cond = []

                                            for i in range(len(back_conds)):
                                                name_back, back_n = Scrape_name_and_click_through(i)

                                                if name_back in list_back_cond:
                                                    continue
                                                else:
                                                    list_back_cond.append(name_back)

                                                back_n.click()

                                                time.sleep(1.5)

                                                price = Get_price()

                                                Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                                Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                                counter += 1

                                                Go_back()

                                            Go_back()

                                        Go_back()

                                    if phone_name == 'Galaxy A50 (2019)':
                                        # Going back for memory_gb options
                                        Go_back()
                                    else:
                                        pass

                                if phone_name == 'Galaxy S10e':
                                    # Going back for memory_gb options
                                    Go_back()
                                else:
                                    pass

                                # Going back for memory size options
                                Go_back()
                                time.sleep(0.5)

                                # Going back for color options
                                Go_back()

                            # Checking whether the phone condition is Poor or Defekt in order to skip
                            elif gen_name == 'Mangelhaft / Defekt':
                                print("Passing Condition", gen_name)
                                pass

                        elif phone_name == 'Galaxy A52 5G (2021)' or phone_name == 'Galaxy A52s 5G (2021)':

                            # Checking whether the condition is 'Second hand' because there is different elements
                            if gen_name == 'Gebraucht':
                                dis_conds = Scrape_slider_info()
                                list_dis_cond = []

                                for i in range(len(dis_conds)):
                                    name_dis, dis_n = Scrape_name_and_click_through(i)

                                    if name_dis in list_dis_cond:
                                        continue
                                    else:
                                        list_dis_cond.append(name_dis)

                                    dis_n.click()

                                    time.sleep(1.5)

                                    # Finding the elements for memory sizes of the phone
                                    mem_sizes = Scrape_slider_info()
                                    list_size_name = []

                                    for i in range(len(mem_sizes)):
                                        name_size, size_n = Scrape_name_and_click_through(i)

                                        if name_size in list_size_name:
                                            continue
                                        else:
                                            list_size_name.append(name_size)

                                        size_n.click()

                                        time.sleep(1.5)

                                        # Finding the elements for the frame condition of the phone
                                        frame_conds = Scrape_slider_info()
                                        list_frame_cond = []

                                        for i in range(len(frame_conds)):
                                            name_frame, frame_n = Scrape_name_and_click_through(i)

                                            if name_frame in list_frame_cond:
                                                continue
                                            else:
                                                list_frame_cond.append(name_frame)

                                            frame_n.click()

                                            time.sleep(1.5)

                                            # Finding the elements for the back condition of the phone
                                            back_conds = Scrape_slider_info()
                                            list_back_cond = []

                                            for i in range(len(back_conds)):
                                                name_back, back_n = Scrape_name_and_click_through(i)

                                                if name_back in list_back_cond:
                                                    continue
                                                else:
                                                    list_back_cond.append(name_back)

                                                back_n.click()

                                                time.sleep(1.5)

                                                # Finding the elements for the colors of the phone
                                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                                # Clicking to continue to next combination
                                                colors[0].click()
                                                time.sleep(1.5)

                                                price = Get_price()

                                                Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                                Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                                counter+=1

                                                Go_back()
                                                time.sleep(0.5)

                                                Go_back()

                                            Go_back()

                                        Go_back()

                                    Go_back()

                            # Checking whether the phone condition is Poor or Defekt in order to skip
                            elif gen_name == 'Mangelhaft / Defekt':
                                print("Passing Condition", gen_name)
                                pass

                        elif phone_name == 'Galaxy A52 5G (2021)' or phone_name == 'Galaxy A72 5G (2021)' or phone_name == 'Galaxy A32 4G (2021)' or phone_name == 'Galaxy A52 4G (2021)':
                            name_size = '128GB'
                            # Checking whether the condition is 'Second hand' because there is different elements
                            if gen_name == 'Gebraucht':
                                if phone_name == 'Galaxy A32 4G (2021)':
                                    memory_gb=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                    memory = memory_gb[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                    # Clicking to continue to next combination
                                    memory[0].click()
                                    time.sleep(1)

                                    # Finding the elements for the colors of the phone
                                    colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                    colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                    # Clicking to continue to next combination
                                    colors[0].click()
                                    time.sleep(1.5)

                                dis_conds = Scrape_slider_info()
                                list_dis_cond = []

                                for i in range(len(dis_conds)):
                                    name_dis, dis_n = Scrape_name_and_click_through(i)

                                    if name_dis in list_dis_cond:
                                        continue
                                    else:
                                        list_dis_cond.append(name_dis)

                                    dis_n.click()

                                    time.sleep(2.5)

                                    if phone_name == 'Galaxy A32 4G (2021)':
                                        pass
                                    else:
                                        time.sleep(2)
                                        if 'Speichergrösse' in driver.find_elements_by_xpath("//div[@class='question assistant']")[-1].text:
                                            memory_gb=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                            memory = memory_gb[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                            # Clicking to continue to next combination
                                            memory[0].click()
                                            time.sleep(1)

                                    # Finding the elements for the frame condition of the phone
                                    frame_conds = Scrape_slider_info()
                                    list_frame_cond = []

                                    for i in range(len(frame_conds)):
                                        name_frame, frame_n = Scrape_name_and_click_through(i)

                                        if name_frame in list_frame_cond:
                                            continue
                                        else:
                                            list_frame_cond.append(name_frame)

                                        frame_n.click()

                                        time.sleep(1.5)

                                        # Finding the elements for the back condition of the phone
                                        back_conds = Scrape_slider_info()
                                        list_back_cond = []

                                        for i in range(len(back_conds)):
                                            name_back, back_n = Scrape_name_and_click_through(i)

                                            if name_back in list_back_cond:
                                                continue
                                            else:
                                                list_back_cond.append(name_back)

                                            back_n.click()

                                            time.sleep(1.5)

                                            if phone_name == 'Galaxy A32 4G (2021)':
                                                pass
                                            else:
                                                # Finding the elements for the colors of the phone
                                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                                # Clicking to continue to next combination
                                                colors[0].click()
                                                time.sleep(1.5)

                                            price = Get_price()

                                            Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                            Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                            counter += 1

                                            if phone_name == 'Galaxy A32 4G (2021)':
                                                pass
                                            else:
                                                Go_back()
                                                time.sleep(0.5)

                                            Go_back()

                                        Go_back()

                                    Go_back()
                                    time.sleep(0.5)

                                    if phone_name == 'Galaxy A32 4G (2021)':
                                        pass
                                    else:
                                        Go_back()

                                if phone_name == 'Galaxy A32 4G (2021)':
                                    Go_back()
                                    time.sleep(0.5)

                                    Go_back()

                            # Checking whether the phone condition is Poor or Defekt in order to skip
                            elif gen_name == 'Mangelhaft / Defekt':
                                print("Passing Condition", gen_name)
                                pass

                            else:
                                time.sleep(1)
                                # Finding the elements for the colors of the phone
                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                colors[0].click()
                                time.sleep(1.5)

                                price = Get_price()

                                name_dis = np.NaN
                                name_frame = np.NaN
                                name_back = np.NaN

                                Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                counter += 1

                                Go_back()
                                time.sleep(0.5)

                                Go_back()

                        else:
                            # Finding the elements for memory sizes of the phone
                            mem_sizes = Scrape_slider_info()
                            list_size_name = []

                            for i in range(len(mem_sizes)):
                                name_size, size_n = Scrape_name_and_click_through(i)

                                if name_size in list_size_name:
                                    continue
                                else:
                                    list_size_name.append(name_size)

                                size_n.click()

                                time.sleep(1.5)

                                # Finding the elements for the colors of the phone
                                colors_div=driver.find_elements_by_xpath("//div[@class='wrapUserAnswers']")
                                colors = colors_div[-1].find_elements_by_xpath('//a[@class="verkaufenPageAnswersBeta"]')

                                # Clicking to continue to next combination
                                colors[0].click()
                                time.sleep(1.5)

                                # Checking whether the condition is 'Second hand' because there is different elements
                                if gen_name == 'Gebraucht':

                                    dis_conds = Scrape_slider_info()
                                    list_dis_cond = []

                                    for i in range(len(dis_conds)):
                                        name_dis, dis_n = Scrape_name_and_click_through(i)

                                        if name_dis in list_dis_cond:
                                            continue
                                        else:
                                            list_dis_cond.append(name_dis)

                                        dis_n.click()

                                        time.sleep(1.5)

                                        # Finding the elements for the frame condition of the phone
                                        frame_conds = Scrape_slider_info()
                                        list_frame_cond = []

                                        for i in range(len(frame_conds)):
                                            name_frame, frame_n = Scrape_name_and_click_through(i)

                                            if name_frame in list_frame_cond:
                                                continue
                                            else:
                                                list_frame_cond.append(name_frame)

                                            frame_n.click()

                                            time.sleep(1.5)

                                            # Finding the elements for the back condition of the phone
                                            back_conds = Scrape_slider_info()
                                            list_back_cond = []

                                            for i in range(len(back_conds)):
                                                name_back, back_n = Scrape_name_and_click_through(i)

                                                if name_back in list_back_cond:
                                                    continue
                                                else:
                                                    list_back_cond.append(name_back)

                                                back_n.click()

                                                time.sleep(1.5)

                                                price = Get_price()

                                                Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                                Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                                counter += 1

                                                Go_back()

                                            Go_back()

                                        Go_back()

                                # Checking whether the phone condition is Poor or Defekt in order to skip
                                elif gen_name == 'Mangelhaft / Defekt':
                                    print("Passing Condition", gen_name)
                                    pass

                                # If the phone condition is New (Welded) then execute this block
                                else:
                                    sleep(1.5)

                                    price = Get_price()

                                    name_dis = np.NaN
                                    name_frame = np.NaN
                                    name_back = np.NaN

                                    Print_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price)

                                    Store_info(phone_url, phone_name, gen_name, name_size, name_dis, name_frame, name_back, price, counter)
                                    counter += 1

                                # Going back for memory size options
                                Go_back()
                                time.sleep(0.5)

                                # Going back for color options
                                Go_back()

                    # Going back for phone condition options
                    Go_back()
 
    except:
        url = phone_url
        # Saving the file so we can concatenate after
        excel_name = 'stopped_at_{}.xlsx'.format(phone_name)
        excel.save(excel_name)
        
        driver.close()
        
        #adding random user agent so the page won't recognize that is a bot
        ua = UserAgent()
        userAgent = ua.random

        opt = webdriver.ChromeOptions()
        opt.add_argument(f'user-agent={userAgent}')
        opt.add_argument("--start-maximized")
        opt.add_argument('--disable-site-isolation-trials')
#         opt.add_argument("--headless")

        # Creating the driver object (this will work for every chrome version).
        chromedriver_autoinstaller.install()
        driver = webdriver.Chrome(options=opt)
        
        index_ = index_at
        counter_fails = 1
        
        #call the scrapper to run
        Scraping(driver, counter_fails = counter_fails, index=index_)
        
    # Making new directory to store the export data
    try:
        path = path_
        os.mkdir(path)
    except:
        pass

    # Saving the file in xlsx table
    mask = "%Y-%m-%d-%H-%M-%S"
    date_now = dt.now().strftime(mask)

    name_fails = name_fails_
    name_succs = name_success_

    if counter_fails > 0:
        excel.save(name_fails)  
    else:
        excel.save(name_succs) 

    # take the name of all excel files in the current folder
    file_list = glob.glob("*.xlsx")

    if name_succs in file_list:
        shutil.move(name_succs, path + '/' + name_succs)
        pass
    else:
        # list of excel files we want to merge.
        # pd.read_excel(file_path) reads the excel
        # data into pandas dataframe.
        excl_list = []

        for file in file_list:
            excl_list.append(pd.read_excel(file))

        # create a new dataframe to store the
        # merged excel file.
        excl_merged = pd.DataFrame()

        for excl_file in excl_list:

            # appends the data into the excl_merged
            # dataframe.
            excl_merged = excl_merged.append(excl_file, ignore_index=True)

        # exports the dataframe into excel file with
        # specified name.

        excl_merged = excl_merged.drop_duplicates()

        mask = "%Y-%m-%d-%H-%M-%S"
        date_now = dt.now().strftime(mask)
        name = name_

        # Saving the file in the new directory with specified date
        excl_merged.to_excel(path + '/' + name, index=False)

        for item in file_list:
            os.remove(item)

    # Printing the time of execution
    print('Execution time {:.2f} minutes'.format((time.time()-start_time)/60))
                
ua = UserAgent()
userAgent = ua.random

opt = webdriver.ChromeOptions()
opt.add_argument(f'user-agent={userAgent}')
opt.add_argument("--start-maximized")
opt.add_argument('--disable-site-isolation-trials')
# opt.add_argument("--headless")

# Creating the driver object (this will work for every chrome version).
chromedriver_autoinstaller.install()
driver = webdriver.Chrome(options=opt)

#call the scrapper to run
Scraping(driver)